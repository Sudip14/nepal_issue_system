import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
import io
from rest_framework import viewsets, permissions, status, generics
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Q
from django.utils import timezone
from .models import User, Department, Issue, Vote, IssueUpdate, Notification, Comment, IssueImage
from .serializers import (
    RegisterSerializer, UserSerializer, DepartmentSerializer,
    IssueListSerializer, IssueDetailSerializer, IssueCreateSerializer,
    IssueUpdateSerializer, NotificationSerializer,
    CommentSerializer, CommentCreateSerializer, IssueImageSerializer
)


class IsAuthorityOrAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.user_type in ['authority', 'admin', 'super_admin']


# ── Auth ──────────────────────────────────────────────────────────────────────

class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]


class MeView(generics.RetrieveUpdateAPIView):
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user


# ── Issues ────────────────────────────────────────────────────────────────────

class IssueViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['title', 'description', 'address', 'ward_number']
    filterset_fields = ['category', 'status', 'severity']
    ordering_fields = ['priority_score', 'created_at', 'vote_count']
    ordering = ['-priority_score']

    def get_queryset(self):
        qs = Issue.objects.filter(is_spam=False).select_related(
            'reporter', 'assigned_department'
        ).prefetch_related('votes', 'updates', 'comments', 'images')

        user = self.request.user
        if user.is_authenticated and user.user_type == 'admin':
            qs = qs.filter(district=user.district)

        category = self.request.query_params.get('category')
        status_f = self.request.query_params.get('status')
        severity = self.request.query_params.get('severity')
        ward = self.request.query_params.get('ward')
        search = self.request.query_params.get('search')

        if category:
            qs = qs.filter(category=category)
        if status_f:
            qs = qs.filter(status=status_f)
        if severity:
            qs = qs.filter(severity=severity)
        if ward:
            qs = qs.filter(ward_number=ward)
        if search:
            qs = qs.filter(Q(title__icontains=search) | Q(description__icontains=search))

        return qs

    def get_serializer_class(self):
        if self.action == 'create':
            return IssueCreateSerializer
        if self.action == 'retrieve':
            return IssueDetailSerializer
        return IssueListSerializer

    def perform_create(self, serializer):
        issue = serializer.save(
            reporter=self.request.user,
            district=self.request.user.district
        )
        self._auto_assign_department(issue)
        issue.calculate_priority_score()

        # 🆕 HANDLE MULTIPLE IMAGES
        images = self.request.FILES.getlist('images')
        for index, image in enumerate(images[:5]):  # Max 5 images
            IssueImage.objects.create(
                issue=issue,
                image=image,
                order=index
            )

    def _auto_assign_department(self, issue):
        category_map = {
            'road': 'Roads',
            'water': 'Water Supply',
            'power': 'Electricity & Power',
            'waste': 'Waste Management',
            'sanitation': 'Waste Management',
            'drainage': 'Water Supply',
        }
        dept_category = category_map.get(issue.category)
        if dept_category:
            dept = Department.objects.filter(issue_category=dept_category).first()
            if dept:
                issue.assigned_department = dept
                issue.status = 'assigned'
                issue.save(update_fields=['assigned_department', 'status'])

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def vote(self, request, pk=None):
        issue = self.get_object()
        vote, created = Vote.objects.get_or_create(issue=issue, voter=request.user)
        if not created:
            vote.delete()
            issue.calculate_priority_score()
            return Response({'voted': False, 'vote_count': issue.votes.count()})
        issue.calculate_priority_score()
        return Response({'voted': True, 'vote_count': issue.votes.count()}, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['patch'], permission_classes=[IsAuthorityOrAdmin])
    def update_status(self, request, pk=None):
        issue = self.get_object()
        new_status = request.data.get('status')
        if new_status not in dict(Issue.STATUS_CHOICES):
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)
        old_status = issue.status
        issue.status = new_status
        if new_status == 'resolved':
            issue.resolved_at = timezone.now()
        issue.save()
        IssueUpdate.objects.create(
            issue=issue,
            updated_by=request.user,
            update_type='status_change',
            message=request.data.get('message', f'Status updated to {new_status}'),
            old_status=old_status,
            new_status=new_status,
            contractor_name=request.data.get('contractor_name', ''),
            estimated_completion_days=request.data.get('estimated_completion_days'),
            inspection_date=request.data.get('inspection_date'),
        )
        if issue.reporter:
            Notification.objects.create(
                user=issue.reporter,
                issue=issue,
                channel='app',
                message=f'Your issue "{issue.title}" status changed from {old_status} to {new_status}.',
                is_sent=True,
            )
        return Response(IssueDetailSerializer(issue, context={'request': request}).data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        qs = Issue.objects.filter(is_spam=False)
        if request.user.is_authenticated and request.user.user_type == 'admin':
            qs = qs.filter(district=request.user.district)
        return Response({
            'total': qs.count(),
            'pending': qs.filter(status='pending').count(),
            'under_review': qs.filter(status='under_review').count(),
            'assigned': qs.filter(status='assigned').count(),
            'in_progress': qs.filter(status='in_progress').count(),
            'resolved': qs.filter(status='resolved').count(),
            'by_category': list(qs.values('category').annotate(count=Count('id'))),
            'by_severity': list(qs.values('severity').annotate(count=Count('id'))),
        })

    @action(detail=False, methods=['get'])
    def heatmap(self, request):
        qs = Issue.objects.filter(is_spam=False)
        if request.user.is_authenticated and request.user.user_type == 'admin':
            qs = qs.filter(district=request.user.district)
        data = qs.values('ward_number').annotate(
            total=Count('id'),
            pending=Count('id', filter=Q(status='pending')),
            resolved=Count('id', filter=Q(status='resolved')),
        ).order_by('ward_number')
        return Response(list(data))

    # ═══════════════════════════════════════════════════════════════════
    # 🆕 IMAGE GALLERY ENDPOINTS
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['get'], permission_classes=[permissions.AllowAny])
    def gallery(self, request, pk=None):
        """GET /api/issues/{id}/gallery/ — List all images for an issue"""
        issue = self.get_object()
        images = issue.images.all().order_by('order')
        return Response(IssueImageSerializer(images, many=True).data)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def add_image(self, request, pk=None):
        """POST /api/issues/{id}/add_image/ — Add image to existing issue"""
        issue = self.get_object()

        if issue.images.count() >= 5:
            return Response({'error': 'Maximum 5 images allowed'}, status=400)

        image = request.FILES.get('image')
        if not image:
            return Response({'error': 'No image provided'}, status=400)

        issue_image = IssueImage.objects.create(
            issue=issue,
            image=image,
            caption=request.data.get('caption', ''),
            order=issue.images.count()
        )
        return Response(IssueImageSerializer(issue_image).data, status=201)

    # ═══════════════════════════════════════════════════════════════════
    # 🆕 COMMENT ACTIONS
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['get'], permission_classes=[permissions.AllowAny])
    def comments(self, request, pk=None):
        issue = self.get_object()
        top_comments = Comment.objects.filter(
            issue=issue, parent=None, is_deleted=False
        ).order_by('-created_at')

        def build_tree(c):
            data = CommentSerializer(c).data
            replies = Comment.objects.filter(parent=c, is_deleted=False).order_by('created_at')
            data['replies'] = [build_tree(r) for r in replies]
            return data

        return Response([build_tree(c) for c in top_comments])

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticatedOrReadOnly])
    def add_comment(self, request, pk=None):
        issue = self.get_object()

        serializer = CommentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        content = serializer.validated_data['content']
        parent_id = serializer.validated_data.get('parent_id')

        parent = None
        if parent_id:
            parent = get_object_or_404(Comment, id=parent_id, issue=issue)

        comment = Comment.objects.create(
            issue=issue,
            parent=parent,
            user=request.user if request.user.is_authenticated else None,
            user_name=request.user.get_full_name() or request.user.username if request.user.is_authenticated else 'Anonymous',
            user_role=request.user.user_type if request.user.is_authenticated else 'guest',
            content=content
        )

        if issue.reporter and issue.reporter != request.user:
            Notification.objects.create(
                user=issue.reporter,
                issue=issue,
                channel='app',
                message=f"New comment on your issue: {issue.title[:50]}...",
            )

        return Response(CommentSerializer(comment).data, status=status.HTTP_201_CREATED)


# ── Departments ───────────────────────────────────────────────────────────────

class DepartmentViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [permissions.AllowAny]


# ── Notifications ─────────────────────────────────────────────────────────────

class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')


# ── Export ────────────────────────────────────────────────────────────────────

class ExportExcelView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            issues = Issue.objects.filter(is_spam=False).select_related('reporter', 'assigned_department').order_by('-created_at')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Issues Report"

            header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)

            headers = ['ID', 'Title', 'Category', 'Severity', 'Status', 'Ward', 'Address', 'Reporter', 'Votes', 'Priority Score', 'Created At']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row, issue in enumerate(issues, 2):
                ws.cell(row=row, column=1, value=issue.id)
                ws.cell(row=row, column=2, value=issue.title)
                ws.cell(row=row, column=3, value=issue.category)
                ws.cell(row=row, column=4, value=issue.get_severity_display())
                ws.cell(row=row, column=5, value=issue.status.replace('_', ' '))
                ws.cell(row=row, column=6, value=str(issue.ward_number))
                ws.cell(row=row, column=7, value=issue.address)
                ws.cell(row=row, column=8, value=f"{issue.reporter.first_name} {issue.reporter.last_name}".strip() if issue.reporter else '')
                ws.cell(row=row, column=9, value=issue.votes.count())
                ws.cell(row=row, column=10, value=round(float(issue.priority_score or 0), 2))
                ws.cell(row=row, column=11, value=issue.created_at.strftime('%Y-%m-%d %H:%M') if issue.created_at else '')

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="issues_report.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        except Exception as e:
            return HttpResponse(str(e), status=500)


class ExportPDFView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            issues = Issue.objects.filter(is_spam=False).select_related('reporter').order_by('-created_at')

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("Nepal Issue Reporting System — Issues Report", styles['Title']))
            elements.append(Spacer(1, 12))

            data = [['ID', 'Title', 'Category', 'Status', 'Ward', 'Severity', 'Votes', 'Priority', 'Created']]
            for issue in issues:
                data.append([
                    str(issue.id),
                    issue.title[:35] + '...' if len(issue.title) > 35 else issue.title,
                    issue.category,
                    issue.status.replace('_', ' '),
                    str(issue.ward_number),
                    issue.get_severity_display(),
                    str(issue.votes.count()),
                    str(round(float(issue.priority_score or 0), 1)),
                    issue.created_at.strftime('%Y-%m-%d') if issue.created_at else '',
                ])

            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(table)
            doc.build(elements)

            buffer.seek(0)
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = 'attachment; filename="issues_report.pdf"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        except Exception as e:
            return HttpResponse(str(e), status=500)